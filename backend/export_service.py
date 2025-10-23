"""
Servicio de Exportación/Importación de Datos
Genera plantillas Excel y procesa importaciones
"""
from datetime import datetime
from typing import List, Dict, Any
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from motor.motor_asyncio import AsyncIOMotorDatabase

# Colores para la plantilla
COLOR_HEADER = "4472C4"  # Azul
COLOR_REQUIRED = "FFF2CC"  # Amarillo
COLOR_OPTIONAL = "E7E6E6"  # Gris claro
COLOR_EXAMPLE = "D9E1F2"  # Azul claro

def create_excel_template() -> io.BytesIO:
    """
    Crea plantilla Excel con todas las hojas y validaciones
    """
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # ==== HOJA 1: INSTRUCCIONES ====
    ws_instructions = wb.create_sheet("📋 INSTRUCCIONES", 0)
    ws_instructions.merge_cells('A1:F1')
    title_cell = ws_instructions['A1']
    title_cell.value = "🏊 PLANTILLA DE IMPORTACIÓN - ESPACIOS CON PISCINA"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_instructions.row_dimensions[1].height = 30
    
    instructions = [
        ["", ""],
        ["📖 CÓMO USAR ESTA PLANTILLA:", ""],
        ["", ""],
        ["1️⃣", "Completa cada hoja según tus datos históricos"],
        ["2️⃣", "Las celdas AMARILLAS son obligatorias"],
        ["3️⃣", "Las celdas GRISES son opcionales"],
        ["4️⃣", "Usa las listas desplegables donde aparezcan"],
        ["5️⃣", "Formatos de fecha: DD/MM/YYYY (ej: 15/10/2025)"],
        ["6️⃣", "Números sin símbolos: 15000 (no RD$15,000)"],
        ["7️⃣", "Guarda el archivo cuando termines"],
        ["8️⃣", "Importa desde la app: Botón 'Importar desde Excel'"],
        ["", ""],
        ["⚠️ IMPORTANTE:", ""],
        ["", ""],
        ["• NO cambies los nombres de las columnas", ""],
        ["• NO elimines las hojas", ""],
        ["• Puedes agregar más filas según necesites", ""],
        ["• La primera fila con datos es un EJEMPLO (puedes eliminarla)", ""],
        ["", ""],
        ["🎨 CÓDIGO DE COLORES:", ""],
        ["", ""],
        ["🟨", "Amarillo = Campo Obligatorio"],
        ["🔲", "Gris = Campo Opcional"],
        ["🟦", "Azul = Fila de Ejemplo (eliminar antes de importar)"],
        ["", ""],
        ["📧 SOPORTE:", ""],
        ["", "Si tienes dudas, contacta al administrador del sistema"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value:
                cell.font = Font(bold=True, size=11)
    
    ws_instructions.column_dimensions['A'].width = 5
    ws_instructions.column_dimensions['B'].width = 60
    
    # ==== HOJA 2: CLIENTES ====
    ws_customers = wb.create_sheet("👥 Clientes")
    
    headers_customers = [
        "Nombre Completo*",
        "Teléfono*",
        "Email",
        "Cédula/Pasaporte/RNC",
        "Dirección"
    ]
    
    # Headers
    for col_idx, header in enumerate(headers_customers, start=1):
        cell = ws_customers.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color según si es obligatorio
        if "*" in header:
            ws_customers.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_REQUIRED, end_color=COLOR_REQUIRED, fill_type="solid")
        else:
            ws_customers.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_OPTIONAL, end_color=COLOR_OPTIONAL, fill_type="solid")
    
    # Ejemplo
    example_customer = ["Juan Pérez", "8091234567", "juan@email.com", "001-1234567-8", "Calle Principal #123"]
    for col_idx, value in enumerate(example_customer, start=1):
        cell = ws_customers.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_EXAMPLE, end_color=COLOR_EXAMPLE, fill_type="solid")
    
    # Anchos de columna
    ws_customers.column_dimensions['A'].width = 25
    ws_customers.column_dimensions['B'].width = 15
    ws_customers.column_dimensions['C'].width = 30
    ws_customers.column_dimensions['D'].width = 20
    ws_customers.column_dimensions['E'].width = 40
    
    # ==== HOJA 3: VILLAS ====
    ws_villas = wb.create_sheet("🏠 Villas")
    
    headers_villas = [
        "Código Villa*",
        "Nombre Villa*",
        "Categoría*",
        "Precio Pasadía DOP*",
        "Precio Amanecida DOP*",
        "Precio Evento DOP*",
        "Precio Propietario DOP*",
        "Descripción",
        "Propietario",
        "Teléfono Propietario",
        "Incluye ITBIS*",
        "Estado*"
    ]
    
    for col_idx, header in enumerate(headers_villas, start=1):
        cell = ws_villas.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        if "*" in header:
            ws_villas.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_REQUIRED, end_color=COLOR_REQUIRED, fill_type="solid")
        else:
            ws_villas.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_OPTIONAL, end_color=COLOR_OPTIONAL, fill_type="solid")
    
    # Ejemplo
    example_villa = ["ECPVSH", "Villa Paraíso", "Premium", "15000", "25000", "30000", "5000", "Villa con piscina grande", "María González", "8099876543", "NO", "Activo"]
    for col_idx, value in enumerate(example_villa, start=1):
        cell = ws_villas.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_EXAMPLE, end_color=COLOR_EXAMPLE, fill_type="solid")
    
    ws_villas.column_dimensions['A'].width = 15
    ws_villas.column_dimensions['B'].width = 25
    ws_villas.column_dimensions['C'].width = 20
    for col in ['D', 'E', 'F', 'G']:
        ws_villas.column_dimensions[col].width = 18
    ws_villas.column_dimensions['H'].width = 30
    ws_villas.column_dimensions['I'].width = 25
    ws_villas.column_dimensions['J'].width = 18
    ws_villas.column_dimensions['K'].width = 15
    ws_villas.column_dimensions['L'].width = 15
    
    # ==== HOJA 4: RESERVACIONES ====
    ws_reservations = wb.create_sheet("🎫 Reservaciones")
    
    headers_reservations = [
        "Número Factura*",
        "Nombre Cliente*",
        "Código Villa*",
        "Tipo Renta*",
        "Fecha Reserva*",
        "Fecha Check-Out",
        "Hora Check-In",
        "Hora Check-Out",
        "Huéspedes*",
        "Precio Base*",
        "Horas Extra",
        "Costo Horas Extra",
        "Descuento",
        "Incluye ITBIS*",
        "Total*",
        "Depósito",
        "Monto Pagado*",
        "Método Pago*",
        "Moneda*",
        "Estado*",
        "Notas"
    ]
    
    for col_idx, header in enumerate(headers_reservations, start=1):
        cell = ws_reservations.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_reservations.row_dimensions[1].height = 30
        
        if "*" in header:
            ws_reservations.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_REQUIRED, end_color=COLOR_REQUIRED, fill_type="solid")
        else:
            ws_reservations.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_OPTIONAL, end_color=COLOR_OPTIONAL, fill_type="solid")
    
    # Ejemplo
    example_reservation = [
        "5815", "Juan Pérez", "ECPVSH", "pasadia", "15/10/2025", "15/10/2025",
        "9:00 AM", "8:00 PM", "10", "15000", "0", "0", "0", "NO",
        "15000", "0", "5000", "efectivo", "DOP", "confirmed", "Cliente regular"
    ]
    for col_idx, value in enumerate(example_reservation, start=1):
        cell = ws_reservations.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_EXAMPLE, end_color=COLOR_EXAMPLE, fill_type="solid")
    
    for col_idx in range(1, len(headers_reservations) + 1):
        ws_reservations.column_dimensions[chr(64 + col_idx)].width = 15
    
    # ==== HOJA 5: GASTOS ====
    ws_expenses = wb.create_sheet("💰 Gastos")
    
    headers_expenses = [
        "Categoría*",
        "Descripción*",
        "Monto*",
        "Moneda*",
        "Fecha*",
        "Estado Pago*",
        "Tipo Gasto*",
        "Tiene Recordatorio",
        "Día Recordatorio",
        "Es Recurrente",
        "Notas"
    ]
    
    for col_idx, header in enumerate(headers_expenses, start=1):
        cell = ws_expenses.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        if "*" in header:
            ws_expenses.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_REQUIRED, end_color=COLOR_REQUIRED, fill_type="solid")
        else:
            ws_expenses.cell(row=2, column=col_idx).fill = PatternFill(start_color=COLOR_OPTIONAL, end_color=COLOR_OPTIONAL, fill_type="solid")
    
    # Ejemplo
    example_expense = ["compromiso", "Pago de luz", "6500", "DOP", "15/10/2025", "pending", "fijo", "SI", "15", "SI", "Pago mensual"]
    for col_idx, value in enumerate(example_expense, start=1):
        cell = ws_expenses.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color=COLOR_EXAMPLE, end_color=COLOR_EXAMPLE, fill_type="solid")
    
    for col in ['A', 'B']:
        ws_expenses.column_dimensions[col].width = 25
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws_expenses.column_dimensions[col].width = 15
    ws_expenses.column_dimensions['H'].width = 18
    ws_expenses.column_dimensions['I'].width = 18
    ws_expenses.column_dimensions['J'].width = 15
    ws_expenses.column_dimensions['K'].width = 30
    
    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


async def export_data_to_excel(db: AsyncIOMotorDatabase, data_type: str) -> io.BytesIO:
    """
    Exporta datos existentes a Excel
    """
    output = io.BytesIO()
    
    if data_type == "customers":
        customers = await db.customers.find({}, {"_id": 0}).to_list(None)
        df = pd.DataFrame(customers)
        if not df.empty:
            df = df[['name', 'phone', 'email', 'identification_document', 'address']]
            df.columns = ['Nombre Completo', 'Teléfono', 'Email', 'Cédula/Pasaporte/RNC', 'Dirección']
        df.to_excel(output, index=False, engine='openpyxl')
    
    elif data_type == "villas":
        villas = await db.villas.find({}, {"_id": 0}).to_list(None)
        df = pd.DataFrame(villas)
        if not df.empty:
            # Mapear columnas según necesidad
            pass
        df.to_excel(output, index=False, engine='openpyxl')
    
    elif data_type == "reservations":
        reservations = await db.reservations.find({}, {"_id": 0}).to_list(None)
        df = pd.DataFrame(reservations)
        df.to_excel(output, index=False, engine='openpyxl')
    
    elif data_type == "expenses":
        expenses = await db.expenses.find({}, {"_id": 0}).to_list(None)
        df = pd.DataFrame(expenses)
        df.to_excel(output, index=False, engine='openpyxl')
    
    output.seek(0)
    return output
